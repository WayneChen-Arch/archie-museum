#!/usr/bin/env python3
"""Sync works.xlsx <-> gallery copy in index.html (assets/).

Usage:
  python3 scripts/sync-works-from-excel.py export
  python3 scripts/sync-works-from-excel.py import
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from opencc import OpenCC

ROOT = Path(__file__).resolve().parents[1]
HTML_PATH = ROOT / "index.html"
XLSX_PATH = ROOT / "works.xlsx"
ASSETS = ROOT / "assets"

HEADERS = [
    "序号",
    "图片文件",
    "画面尺寸",
    "名称",
    "年份",
    "年龄",
    "创作媒介",
    "一句话介绍",
    "创作过程说明",
]

SIZE_TO_LABEL = {
    "a3": "A3",
    "a4": "A4",
    "a5": "A5",
    "tall": "A5加长",
    "wide": "A4",
    "series": "A5",
}

EDITABLE_NOTE = (
    "使用说明：\n"
    "1. 请主要修改「名称」「年份」「创作媒介」「一句话介绍」「创作过程说明」「画面尺寸」。\n"
    "2. 「序号」和「图片文件」请勿随意改动；图片文件对应 assets/ 下不含扩展名的文件名。\n"
    "3. 画面尺寸可用：A3 / A4 / A5 / A4加长 / A5加长。\n"
    "4. 上传覆盖 works.xlsx 后运行：python3 scripts/sync-works-from-excel.py import\n"
    "5. 同步规则：简体写入网页；繁体自动简转繁；英文列若之后扩展可再补。"
)


def s2t(text: str) -> str:
    return OpenCC("s2t").convert(text) if text else ""


def extract_works_span(html: str) -> tuple[int, int]:
    start = html.find("const art = (file) => encodeURI(`assets/${file}`);")
    if start < 0:
        start = html.find("const art = (file) => encodeURI(`drawings/${file}`);")
    if start < 0:
        raise RuntimeError("找不到 art() 定义")
    end = html.find("\n      const mainTrack", start)
    if end < 0:
        raise RuntimeError("找不到 works 后的 mainTrack")
    return start, end


def parse_works(html: str) -> list[dict]:
    start, end = extract_works_span(html)
    block = html[start:end]
    body_m = re.search(r"const works = \[([\s\S]*)\];\s*$", block)
    if not body_m:
        raise RuntimeError("无法解析 works 数组")
    body = body_m.group(1).strip()
    if body.endswith(","):
        body = body[:-1]

    objs: list[str] = []
    depth = 0
    start_i = None
    for i, ch in enumerate(body):
        if ch == "{":
            if depth == 0:
                start_i = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start_i is not None:
                objs.append(body[start_i : i + 1])
                start_i = None

    works = []
    for obj in objs:
        works.append(
            {
                "titles": _pick_i18n(obj, "titles"),
                "year": _pick_str(obj, "year"),
                "materials": _pick_i18n(obj, "materials"),
                "image_file": _pick_art(obj, "image"),
                "size": _pick_str(obj, "size"),
                "notes": _pick_i18n(obj, "notes"),
                "process_file": _pick_art(obj, "processImage"),
                "processCaptions": _pick_i18n(obj, "processCaptions") or {},
                "series": _pick_str(obj, "series"),
                "panel": _pick_str(obj, "panel"),
            }
        )
    return works


def _pick_str(obj: str, key: str) -> str:
    m = re.search(rf'{key}:\s*"((?:\\.|[^"\\])*)"', obj)
    return _unescape(m.group(1)) if m else ""


def _pick_art(obj: str, key: str) -> str:
    m = re.search(rf'{key}:\s*art\("((?:\\.|[^"\\])*)"\)', obj)
    return _unescape(m.group(1)) if m else ""


def _pick_i18n(obj: str, key: str) -> dict:
    m = re.search(rf"{key}:\s*\{{([^{{}}]+)\}}", obj, re.S)
    if not m:
        return {}
    body = m.group(1)
    out = {}
    for lang in ("zh-Hans", "zh-Hant", "en"):
        lm = re.search(rf'(?:"{lang}"|{re.escape(lang)}):\s*"((?:\\.|[^"\\])*)"', body)
        if lm:
            out[lang] = _unescape(lm.group(1))
    return out


def _unescape(s: str) -> str:
    return s.replace("\\n", "\n").replace("\\r", "").replace('\\"', '"').replace("\\\\", "\\")


def _escape_js(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "")


def _i18n_literal(zh_hans: str, zh_hant: str, en: str) -> str:
    return (
        '{ "zh-Hans": "'
        + _escape_js(zh_hans)
        + '", "zh-Hant": "'
        + _escape_js(zh_hant)
        + '", en: "'
        + _escape_js(en)
        + '" }'
    )


def export_xlsx(works: list[dict]) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "使用说明"
    guide["A1"] = EDITABLE_NOTE
    guide["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 96
    guide.row_dimensions[1].height = 140

    ws = wb.create_sheet("作品资料", 0)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A5560")
    edit_fill = PatternFill("solid", fgColor="FFF7E6")
    lock_fill = PatternFill("solid", fgColor="F0F0F0")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill

    for idx, work in enumerate(works, 1):
        stem = Path(work["image_file"]).stem
        row = [
            idx,
            stem,
            SIZE_TO_LABEL.get(work["size"], work["size"].upper()),
            work["titles"].get("zh-Hans", ""),
            work["year"],
            "",
            work["materials"].get("zh-Hans", ""),
            work["notes"].get("zh-Hans", "") or "无",
            work["processCaptions"].get("zh-Hans", ""),
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(idx + 1, col, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = lock_fill if col in (1, 2) else edit_fill

    widths = [6, 28, 10, 22, 12, 8, 14, 48, 36]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"已导出 {len(works)} 件作品到 {XLSX_PATH}")


def map_size_label(raw: str) -> str:
    raw = (raw or "").strip()
    if "加长" in raw or "加長" in raw:
        return "tall"
    up = raw.upper()
    if up.startswith("A3"):
        return "a3"
    if up.startswith("A5"):
        return "a5"
    if up.startswith("A4"):
        return "a4"
    return "a4"


def find_asset(stem: str) -> str:
    cands = [
        p.name
        for p in ASSETS.iterdir()
        if p.name != "4360.png"
        and p.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}
        and p.stem == stem
    ]
    if not cands:
        raise FileNotFoundError(f"assets 中找不到：{stem}")
    non_proc = [n for n in cands if "创作过程" not in n]
    return sorted(non_proc or cands)[0]


def find_process(stem: str) -> str:
    for p in ASSETS.iterdir():
        if p.stem.startswith(stem) and "创作过程" in p.name:
            return p.name
    return ""


def load_excel_rows() -> list[dict]:
    wb = load_workbook(XLSX_PATH)
    ws = wb["作品资料"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers[:9] != HEADERS:
        raise RuntimeError(f"表头不匹配。期望 {HEADERS}，实际 {headers[:9]}")
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values or not values[1]:
            continue
        year = values[4]
        if isinstance(year, datetime):
            year_s = f"{year.year}.{year.month}.{year.day}"
        else:
            year_s = "" if year is None else str(year).strip()
        note = "" if values[7] is None else str(values[7]).strip()
        process = "" if values[8] is None else str(values[8]).strip()
        rows.append(
            {
                "stem": str(values[1]).strip(),
                "paper": str(values[2]).strip() if values[2] else "",
                "title": str(values[3]).strip() if values[3] else str(values[1]).strip(),
                "year": year_s,
                "medium": str(values[6]).strip() if values[6] else "",
                "note": "" if note in ("无", "None") else note,
                "process": "" if process in ("无", "None") else process,
            }
        )
    return rows


SERIES_RULES = {
    "地球 I": ("earth-duo", "I / II"),
    "地球 II": ("earth-duo", "II / II"),
    "人物 I": ("portrait-duo", "I / II"),
    "人物 II": ("portrait-duo", "II / II"),
    "感谢信 I": ("thank-you-teachers", "I / III"),
    "感谢信 II": ("thank-you-teachers", "II / III"),
    "感谢信 III": ("thank-you-teachers", "III / III"),
}


def rebuild_work_object(old: dict | None, row: dict) -> str:
    title_zh = row["title"]
    material_zh = row["medium"]
    note_zh = row["note"]
    process_zh = row["process"] or note_zh
    title_en = (old or {}).get("titles", {}).get("en") or title_zh
    material_en = (old or {}).get("materials", {}).get("en") or material_zh
    note_en = (old or {}).get("notes", {}).get("en") or note_zh
    process_en = (old or {}).get("processCaptions", {}).get("en") or process_zh
    year = row["year"] or ((old or {}).get("year") or "")
    image_file = find_asset(row["stem"])
    size = map_size_label(row["paper"]) or ((old or {}).get("size") or "a4")
    series_info = SERIES_RULES.get(title_zh)
    if not series_info and old:
        if old.get("series"):
            series_info = (old["series"], old.get("panel") or "")

    lines = [
        "        {",
        f"          titles: {_i18n_literal(title_zh, s2t(title_zh), title_en)},",
        f'          year: "{_escape_js(year)}",',
        f"          materials: {_i18n_literal(material_zh, s2t(material_zh), material_en)},",
        f'          image: art("{_escape_js(image_file)}"),',
        f'          size: "{_escape_js(size)}",',
    ]
    if series_info:
        sid, panel = series_info
        lines.append(f'          series: "{_escape_js(sid)}",')
        if panel:
            lines.append(f'          panel: "{_escape_js(panel)}",')
    proc_file = find_process(row["stem"]) or (old or {}).get("process_file") or ""
    if proc_file:
        lines.append(f'          processImage: art("{_escape_js(proc_file)}"),')
        lines.append("          processCaptions: {")
        lines.append(f'            "zh-Hans": "{_escape_js(process_zh)}",')
        lines.append(f'            "zh-Hant": "{_escape_js(s2t(process_zh))}",')
        lines.append(f'            en: "{_escape_js(process_en)}"')
        lines.append("          },")
    lines.append(f"          notes: {_i18n_literal(note_zh, s2t(note_zh), note_en)}")
    lines.append("        }")
    return "\n".join(lines)


def import_xlsx() -> None:
    html = HTML_PATH.read_text(encoding="utf-8")
    old_works = parse_works(html)
    by_stem = {Path(w["image_file"]).stem: w for w in old_works}
    # also map by current titles for rematch after renames
    by_title = {w["titles"].get("zh-Hans", ""): w for w in old_works}
    rows = load_excel_rows()
    rebuilt = []
    for row in rows:
        old = by_stem.get(row["stem"]) or by_title.get(row["title"])
        rebuilt.append(rebuild_work_object(old, row))

    start, end = extract_works_span(html)
    new_block = (
        "const art = (file) => encodeURI(`assets/${file}`);\n\n      const works = [\n"
        + ",\n".join(rebuilt)
        + "\n      ];"
    )
    HTML_PATH.write_text(html[:start] + new_block + html[end:], encoding="utf-8")
    print(f"已从 {XLSX_PATH} 同步 {len(rebuilt)} 件作品到 {HTML_PATH}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["export", "import"])
    args = parser.parse_args()
    html = HTML_PATH.read_text(encoding="utf-8")
    if args.command == "export":
        export_xlsx(parse_works(html))
    else:
        import_xlsx()
        parse_works(HTML_PATH.read_text(encoding="utf-8"))
        print("二次解析通过。")


if __name__ == "__main__":
    main()
